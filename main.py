#Imports
from typing import Pattern
import yfinance as yf
import openpyxl
import pandas as pd
import numpy as np
from numpy.ma.core import choose
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, AreaChart
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.cell import cell
from copy import copy
from openpyxl.styles.builtins import styles
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice

try:
    #Variables
    stockData = "empty" #Declare stockData outside an if statement
    period = 0
    startDate = 0
    endDate = 0
    validPeriods = {"5d", "1mo", "3mo", "6mo", "1y", "2y", "5y", "10y", "ytd", "max"} #Creates a matrix of valid periods for error handling later
    maxLength = 0
    movingAverageOption = 0
    overBoughtSold = None

    company = input("What company's stocks do you want to look at? (Enter their ticker)\n") #Select company

    while True:
        while True:
            dateOrPeriod = input("Choose whether you want to see results for a specific time range or for a period. r/p\n")  # Choose the data display mode (by period or range)
            if dateOrPeriod == "p": #For period selected
                while True:
                    period = input("Select a period: (Available periods are 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max)\n") #Select period
                    if period in validPeriods:
                        stockData = yf.download(company, period=period, auto_adjust=True) #Grab financial data
                        break
                    else:
                        print("Invalid input, please try again\n")  # loops the code if an invalid statement is given
                break
            elif dateOrPeriod == "r": #For range selected
                startDate = input("Select a start date (format: Year-Month-Date)\n")
                endDate = input("Select an end date (format: Year-Month-Date)\n")
                stockData = yf.download(company, start=startDate, end=endDate, auto_adjust=True)
                break
            else:
                print("Invalid input, please try again") #loops the code if an invalid statement is given

        fileName = f"{company}StockData.xlsx"

        if stockData.empty:
            print("Error: No data found for the given stock ticker. Please check the ticker and try again. If the error persists, check your internet connection.\n")
        else:
            print(f"{stockData.head()}\n") #prints the stock data
            stockData["EMA_20"] = stockData["Close"].ewm(span=20, adjust=False).mean()  # Uses panda to calculate an exponential moving average (EMA)
            stockData["EMA_50"] = stockData["Close"].ewm(span=50, adjust=False).mean()
            stockData["EMA_200"] = stockData["Close"].ewm(span=200, adjust=False).mean()
            stockData["STD"] = stockData["Close"].rolling(window=20).std()
            stockData["Upper Band"] = stockData["EMA_20"] + (2 * stockData["STD"])
            stockData["Lower Band"] = stockData["EMA_20"] - (2 * stockData["STD"])

            delta = stockData["Close"].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            stockData["RSI"] = 100 - (100 / (1 + rs))

            stockData.to_excel(fileName, engine='openpyxl') #Send the data to an Excel sheet

        wb = load_workbook(fileName) #establish the workbook without printing formulas
        ws = wb.active #make the worksheet the active workbook

        for col in ws.columns:
            colLetter = col[0].column_letter
            for cell in col:
                if isinstance(cell.value, (int, float)): #Reformats decimal points if the cell contains a number
                    cell.number_format = "0.00"
                    maxLength = max(maxLength, len(f"{cell.value:.2f}")) #Track the max length of formatted numbers
                else:
                    maxLength = max(maxLength, len(str(cell.value))) #Track the max length of non-numeric values

            ws.column_dimensions[colLetter].width = maxLength #Extends the cells to fit the contents

        ws["J1"] = "Change" #Adds a change column
        ws["J1"].font = copy(ws["F1"].font)
        ws["J1"].alignment = copy(ws["F1"].alignment)
        ws["J1"].fill = copy(ws["F1"].fill)
        ws["G1"].border = copy(ws["F1"].border)
        ws.column_dimensions["J"].width = maxLength + 1

        if dateOrPeriod == "p":
            ws["B3"] = f"Data for the past {period}"
            ws["B3"].font = copy(ws["F1"].font)
            ws["B3"].alignment = copy(ws["F1"].alignment)
            ws["B3"].fill = copy(ws["F1"].fill)
            ws["B3"].border = copy(ws["F1"].border)
        else:
            ws["B3"] = f"Data from {startDate} to {endDate}"
            ws["B3"].font = copy(ws["F1"].font)
            ws["B3"].alignment = copy(ws["F1"].alignment)
            ws["B3"].fill = copy(ws["F1"].fill)
            ws["B3"].border = copy(ws["F1"].border)

        green = PatternFill(start_color="568203", end_color="568203", fill_type="solid")
        red = PatternFill(start_color="FF2800", end_color="FF2800", fill_type="solid")
        redFont = Font(color="FF2800", bold=True)
        greenFont = Font(color="568203", bold=True)

        column = "B"

        ws.conditional_formatting.add( #Adds conditional formatting
            f"{column}4:{column}{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["B3"], stopIfTrue=False, font=greenFont)
        )
        ws.conditional_formatting.add(
            f"{column}4:{column}{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["B3"], stopIfTrue=False, font=redFont)
        )

        ws["G5"].value = "=B5-B4" #Calculates the change in stock price
        i = 5 #Starts from cell G5
        while i <= ws.max_row: #Loops through each cell in the column
            ws[f"J{i}"] = f"=B{i}-B{i - 1}"
            i += 1

        column2 = "J" #Reformatting values again here because data written by us is formatted separately from that of the program
        for row in ws[f"{column2}2:{column2}{ws.max_row}"]:
            for cell in row:
                cell.number_format = "0.00"

        ws.conditional_formatting.add( #Add more to non-pandas data
            f"{column2}2:{column2}{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, font=greenFont)
        )
        ws.conditional_formatting.add(
            f"{column2}2:{column2}{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, font=redFont)
        )

        chart = LineChart() #Creates the line chart
        chart.smooth = False
        chart.style = 12
        chart.height = 23
        chart.width = 30
        chart.title = f"Stock prices for {company}"
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Price"

        data = Reference(ws, min_col=2, min_row=4, max_row=ws.max_row) #Takes the data needed for the chart from the inputted data for the data
        movingAverageReference20 = Reference(ws, min_col=7, min_row=4, max_row=ws.max_row) #Creates a chart reference for the moving averages from before
        movingAverageReference50 = Reference(ws, min_col=8, min_row=4, max_row=ws.max_row)
        movingAverageReference200 = Reference(ws, min_col=9, min_row=4, max_row=ws.max_row)
        upperBandReference = Reference(ws, min_col=11, min_row=4, max_row=ws.max_row)
        lowerBandReference = Reference(ws, min_col=12, min_row=4, max_row=ws.max_row)
        dates = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row) #same for the dates

        chart.add_data(data, titles_from_data=False) #Inputs the data
        chart.add_data(movingAverageReference20, titles_from_data=False) #Inputs the moving averages onto the chart
        chart.add_data(movingAverageReference50, titles_from_data=False)
        chart.add_data(movingAverageReference200, titles_from_data=False)
        chart.add_data(upperBandReference, titles_from_data=False)
        chart.add_data(lowerBandReference, titles_from_data=False)

        chart.series[0].tx = SeriesLabel(v="Stock Close") #Labels the legend
        chart.series[1].tx = SeriesLabel(v=f"20-day EMA")
        chart.series[2].tx = SeriesLabel(v=f"50-day EMA")
        chart.series[3].tx = SeriesLabel(v=f"200-day EMA")
        chart.series[4].tx = SeriesLabel(v="Upper Band")
        chart.series[5].tx = SeriesLabel(v="Lower Band")
        colors = ["000000", "FF2800", "008000", "FFA500", "CF0FFF", "CF0FFF"]

        for i, series in enumerate(chart.series): #Chart lines get visual properties
            series.smooth = False
            series.graphicalProperties.line.solidFill = colors[i]
            series.graphicalProperties.line.width = 17000

        ws["G5"].value = None #Removes an annoying bug

        ws.add_chart(chart, "N16") #Moves the chart to cell M15

        # Adding change summary.
        if dateOrPeriod == "p":
            if period == "5d":
                ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 4})/B{ws.max_row - 4})"
                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
            elif period == "1mo" or period == "3mo" or period == "6mo":
                ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 5})/B{ws.max_row - 5})"
                ws["P3"] = f"=((B{ws.max_row} - B{ws.max_row - 20})/B{ws.max_row - 20})"

                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["O3"] = "=IF(P3 > 0,\"Up\", IF(P3 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
                ws["N3"] = "Last month"
            elif period == "ytd":
                if ws.max_row > 10:
                    ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 5})/B{ws.max_row - 5})"
                if ws.max_row > 27:
                    ws["P3"] = f"=((B{ws.max_row} - B{ws.max_row - 22})/B{ws.max_row - 22})"
                ws["P4"] = f"=((B{ws.max_row} - B4)/B4)"

                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["O3"] = "=IF(P3 > 0,\"Up\", IF(P3 < 0,\"Down\",\"No change\"))"
                ws["O4"] = "=IF(P4 > 0,\"Up\", IF(P4 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
                ws["N3"] = "Last month"
                ws["N4"] = "To date"
            elif period == "1y":
                ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 5})/B{ws.max_row - 5})"
                ws["P3"] = f"=((B{ws.max_row} - B{ws.max_row - 21})/B{ws.max_row - 21})"
                ws["P4"] = f"=((B{ws.max_row} - B{ws.max_row - 249})/B{ws.max_row - 249})"

                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["O3"] = "=IF(P3 > 0,\"Up\", IF(P3 < 0,\"Down\",\"No change\"))"
                ws["O4"] = "=IF(P4 > 0,\"Up\", IF(P4 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
                ws["N3"] = "Last month"
                ws["N4"] = "Past year"
            else:
                ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 5})/B{ws.max_row - 5})"
                ws["P3"] = f"=((B{ws.max_row} - B{ws.max_row - 21})/B{ws.max_row - 21})"
                ws["P4"] = f"=((B{ws.max_row} - B{ws.max_row - 251})/B{ws.max_row - 251})"

                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["O3"] = "=IF(P3 > 0,\"Up\", IF(P3 < 0,\"Down\",\"No change\"))"
                ws["O4"] = "=IF(P4 > 0,\"Up\", IF(P4 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
                ws["N3"] = "Last month"
                ws["N4"] = "Past year"
        else:
            if ws.max_row >= 251:
                ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 5})/B{ws.max_row - 5})"
                ws["P3"] = f"=((B{ws.max_row} - B{ws.max_row - 22})/B{ws.max_row - 22})"
                ws["P4"] = f"=((B{ws.max_row} - B{ws.max_row - 251})/B{ws.max_row - 251})"

                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["O3"] = "=IF(P3 > 0,\"Up\", IF(P3 < 0,\"Down\",\"No change\"))"
                ws["O4"] = "=IF(P4 > 0,\"Up\", IF(P4 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
                ws["N3"] = "Last month"
                ws["N4"] = "Past year"

            elif ws.max_row >= 22:
                ws["P2"] = f"=((B{ws.max_row} - B{ws.max_row - 5})/B{ws.max_row - 5})"
                ws["P3"] = f"=((B{ws.max_row} - B{ws.max_row - 21})/B{ws.max_row - 21})"

                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["O3"] = "=IF(P3 > 0,\"Up\", IF(P3 < 0,\"Down\",\"No change\"))"

                ws["N2"] = "Last week"
                ws["N3"] = "Last month"
            else:
                ws["P2"] = f"=((B{ws.max_row} - B4)/B4)"
                ws["O2"] = "=IF(P2 > 0,\"Up\", IF(P2 < 0,\"Down\",\"No change\"))"
                ws["N2"] = "Last week or less"

        ws["P2"].number_format = '0.00%'
        ws["P3"].number_format = '0.00%'
        ws["P4"].number_format = '0.00%'

        stockData["dailyReturns"] = stockData["Close"].pct_change()
        dailyStandardDeviation = stockData["dailyReturns"].std()

        if ws.max_row >= 256:
            annualVolatility = dailyStandardDeviation * np.sqrt(252)
            ws["P7"] = annualVolatility
            ws["P7"].number_format = '0.00%'
            ws["N7"] = "Annual volatility"

        ws["P6"] = dailyStandardDeviation
        ws["P6"].number_format = '0.00%'
        ws["N6"] = "Daily volatility for the period"

        # Death Cross / Golden Cross detection part remains mostly unchanged
        # (unless you want those M9, N9, M10, N10 cells also moved.)

        deathCrossDetect = None
        goldCrossDetect = None
        detectCountDeath = 0
        detectCountGold = 0
        crossoverFillDeath = PatternFill(start_color="9932cc", end_color="9932cc", fill_type="solid")
        crossoverFillGold = PatternFill(start_color="ffb74a", end_color="ffb74a", fill_type="solid")

        for row in range(8, ws.max_row + 1):
            ema50 = ws[f"H{row}"].value
            ema200 = ws[f"I{row}"].value

            ema50Yesterday = ws[f"H{row - 1}"].value
            ema200Yesterday = ws[f"I{row - 1}"].value
            if None not in (ema50, ema200, ema50Yesterday, ema200Yesterday):
                if ema50Yesterday > ema200Yesterday and ema50 < ema200:
                    deathCrossDetect = ws[f"A{row}"].value
                    print(f"Death cross occurred on {deathCrossDetect}")
                    cellDeath = ws[f"A{row}"]
                    cellDeath.fill = crossoverFillDeath
                    detectCountDeath += 1
                    ws["N9"] = f"Death cross occurred {detectCountDeath} times,"
                    ws["O9"] = f" last on {deathCrossDetect}"
                if ema50Yesterday < ema200Yesterday and ema50 > ema200:
                    goldCrossDetect = ws[f"A{row}"].value
                    print(f"Golden cross occurred on {goldCrossDetect}")
                    cellGold = ws[f"A{row}"]
                    cellGold.fill = crossoverFillGold
                    detectCountGold += 1
                    ws["N10"] = f"Gold cross occurred {detectCountGold} times,"
                    ws["O10"] = f" last on {goldCrossDetect}"

        if not deathCrossDetect:
            print("No death cross occurrences")
            ws["N9"] = "No death cross occurrences"
        if not goldCrossDetect:
            print("No gold cross occurrences")
            ws["N10"] = "No gold cross occurrences"

        trend = None
        closeVal = float(stockData["Close"].iloc[-1].item())
        emaVal = float(stockData["EMA_200"].iloc[-1].item())
        if closeVal > emaVal:
            trend = "Up"
        else:
            trend = "Down"

        ws["N8"] = "Stock is currently trending:"
        ws["O8"] = f"{trend}"

        stockDataForCalc = yf.Ticker(company)

        peRatio = stockDataForCalc.info.get('trailingPE')
        growth = stockDataForCalc.info.get('earningsGrowth')

        try:
            pegRatio = peRatio / growth
            peg = round(pegRatio, 2)
            if dateOrPeriod == "p":
                if peg > 1:
                    ws["N12"] = f"Trailing PEG currently at {peg}, stock may be overvalued"
                elif peg < 0:
                    ws[
                        "N12"] = f"Trailing PEG currently at {peg}, company is currently losing money or has a negative growth rate."
                elif peg < 1:
                    ws["N12"] = f"Trailing PEG currently at {peg}, stock may be undervalued"
                else:
                    ws["N12"] = "Stock is fairly valued"
        except Exception as e:
            print(f"An error occurred: {e}. If you specified an index fund like S&P 500, this is normal.")

        currentRSI = stockData["RSI"].iloc[-1]
        currentRSI = round(currentRSI, 2)

        if currentRSI > 70:
            overBoughtSold = "overbought"
        elif currentRSI < 30:
            overBoughtSold = "oversold"
        else:
            overBoughtSold = "normal"

        ws["N13"] = f"Current RSI is {currentRSI}. Stock is {overBoughtSold}."

        ws["N1"] = "Summary"
        ws["N1"].font = copy(ws["F1"].font)
        ws["N1"].alignment = copy(ws["F1"].alignment)
        ws["N1"].fill = copy(ws["F1"].fill)
        ws.column_dimensions["N"].width = maxLength + 6

        wb.save(fileName) #Saves the current workbook under the above filename
        print(f"Data successfully saved to {fileName}")

        break
except Exception as e:
    print(f"An error occurred: {e}")
